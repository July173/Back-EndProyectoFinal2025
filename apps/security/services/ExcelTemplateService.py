from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse
from io import BytesIO
from apps.security.entity.models import Role
from apps.general.entity.models import Program, Ficha, KnowledgeArea
from datetime import datetime


class ExcelTemplateService:
    """
    Servicio para generar plantillas de Excel para el registro masivo de instructores y aprendices.
    Consulta la base de datos para obtener datos actualizados de roles, programas, fichas, etc.
    """

    def __init__(self):
        self.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.required_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

        self.data_style = {
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _apply_style(self, cell, style_dict):
        """Aplica un estilo específico a una celda"""
        for style_type, style_value in style_dict.items():
            setattr(cell, style_type, style_value)

    def _auto_adjust_columns(self, worksheet):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_data_validation(self, worksheet, column_letter, values, start_row=2, end_row=1000):
        """
        Agrega validación de datos (lista desplegable) a una columna específica
        """
        if not values:
            return
            
        # Crear la validación de datos
        values_str = ','.join([str(v) for v in values])
        dv = DataValidation(
            type="list", 
            formula1=f'"{values_str}"',
            allow_blank=True
        )
        dv.error = 'El valor debe ser seleccionado de la lista'
        dv.errorTitle = 'Valor Inválido'
        dv.prompt = 'Selecciona un valor de la lista desplegable'
        dv.promptTitle = 'Seleccionar Valor'
        
        # Aplicar la validación al rango de celdas
        dv.add(f'{column_letter}{start_row}:{column_letter}{end_row}')
        worksheet.add_data_validation(dv)

    def generate_instructor_template(self):
        """
        Genera la plantilla de Excel para el registro masivo de instructores.
        Incluye datos actualizados de la BD: áreas de conocimiento, tipos de contrato, etc.
        """
        wb = Workbook()
        
        # Hoja principal de datos
        ws_main = wb.active
        ws_main.title = "Instructores"
        
        # Definir encabezados de la plantilla de instructores
        headers = [
            ('PRIMER NOMBRE*', True),
            ('SEGUNDO NOMBRE', False),
            ('PRIMER APELLIDO*', True),
            ('SEGUNDO APELLIDO', False),
            ('TIPO IDENTIFICACIÓN*', True),
            ('NÚMERO IDENTIFICACIÓN*', True),
            ('TELÉFONO*', True),
            ('EMAIL INSTITUCIONAL*', True),
            ('TIPO DE CONTRATO*', True),
            ('FECHA INICIO CONTRATO*', True),
            ('FECHA FIN CONTRATO*', True),
            ('ÁREA DE CONOCIMIENTO*', True),
            ('CONTRASEÑA TEMPORAL*', True)
        ]
        
        # Escribir encabezados
        for col_idx, (header, is_required) in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            style = self.required_style if is_required else self.header_style
            self._apply_style(cell, style)
        
        # Agregar filas de ejemplo con formatos esperados
        example_data = [
            ['Juan', 'Carlos', 'Pérez', 'González', 'CC', '12345678', '3001234567', 
             'juan.perez@sena.edu.co', 'Planta', '2024-01-15', '2024-12-31', 
             'Sistemas', 'TempPass123']
        ]
        
        for row_idx, data_row in enumerate(example_data, 2):
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                self._apply_style(cell, self.data_style)
        
        # Crear hojas auxiliares con datos de la BD
        self._create_knowledge_areas_sheet(wb)
        self._create_contract_types_sheet(wb)
        self._create_identification_types_sheet(wb)
        
        # Agregar validaciones de datos (listas desplegables)
        self._add_instructor_data_validations(ws_main)
        
        # Crear hoja de instrucciones
        self._create_instructor_instructions_sheet(wb)
        
        # Ajustar columnas
        self._auto_adjust_columns(ws_main)
        
        return self._save_workbook_to_response(wb, 'plantilla_instructores.xlsx')

    def generate_aprendiz_template(self):
        """
        Genera la plantilla de Excel para el registro masivo de aprendices.
        Incluye datos actualizados de la BD: programas, fichas, etc.
        """
        wb = Workbook()
        
        # Hoja principal de datos
        ws_main = wb.active
        ws_main.title = "Aprendices"
        
        # Definir encabezados de la plantilla de aprendices
        headers = [
            ('PRIMER NOMBRE*', True),
            ('SEGUNDO NOMBRE', False),
            ('PRIMER APELLIDO*', True),
            ('SEGUNDO APELLIDO', False),
            ('TIPO IDENTIFICACIÓN*', True),
            ('NÚMERO IDENTIFICACIÓN*', True),
            ('TELÉFONO*', True),
            ('EMAIL INSTITUCIONAL*', True),
            ('CÓDIGO PROGRAMA*', True),
            ('NÚMERO FICHA*', True),
            ('CONTRASEÑA TEMPORAL*', True)
        ]
        
        # Escribir encabezados
        for col_idx, (header, is_required) in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            style = self.required_style if is_required else self.header_style
            self._apply_style(cell, style)
        
        # Agregar filas de ejemplo
        example_data = [
            ['María', 'José', 'García', 'López', 'CC', '87654321', '3007654321',
             'maria.garcia@soy.sena.edu.co', 'PROG001', '2567890', 'TempPass456']
        ]
        
        for row_idx, data_row in enumerate(example_data, 2):
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                self._apply_style(cell, self.data_style)
        
        # Crear hojas auxiliares con datos de la BD
        self._create_programs_sheet(wb)
        self._create_fichas_sheet(wb)
        self._create_identification_types_sheet(wb)
        
        # Agregar validaciones de datos (listas desplegables)
        self._add_aprendiz_data_validations(ws_main)
        
        # Crear hoja de instrucciones
        self._create_aprendiz_instructions_sheet(wb)
        
        # Ajustar columnas
        self._auto_adjust_columns(ws_main)
        
        return self._save_workbook_to_response(wb, 'plantilla_aprendices.xlsx')

    def _create_knowledge_areas_sheet(self, workbook):
        """Crea una hoja con las áreas de conocimiento disponibles"""
        ws = workbook.create_sheet("Áreas de Conocimiento")
        
        # Encabezado
        headers = ['ID', 'NOMBRE', 'DESCRIPCIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        
        # Datos de la BD
        knowledge_areas = KnowledgeArea.objects.filter(active=True).order_by('name')
        for row_idx, area in enumerate(knowledge_areas, 2):
            ws.cell(row=row_idx, column=1, value=area.id)
            ws.cell(row=row_idx, column=2, value=area.name)
            ws.cell(row=row_idx, column=3, value=area.description or '')
        
        self._auto_adjust_columns(ws)

    def _create_contract_types_sheet(self, workbook):
        """Crea una hoja con los tipos de contrato disponibles"""
        ws = workbook.create_sheet("Tipos de Contrato")
        
        # Encabezado
        cell = ws.cell(row=1, column=1, value='TIPOS DE CONTRATO DISPONIBLES')
        self._apply_style(cell, self.header_style)
        
        # Tipos de contrato comunes en el SENA
        contract_types = [
            'Planta',
            'Contratista',
            'Temporal',
            'Prestación de Servicios',
            'Cátedra'
        ]
        
        for row_idx, contract_type in enumerate(contract_types, 2):
            ws.cell(row=row_idx, column=1, value=contract_type)
        
        self._auto_adjust_columns(ws)

    def _create_programs_sheet(self, workbook):
        """Crea una hoja con los programas disponibles"""
        ws = workbook.create_sheet("Programas")
        
        # Encabezado
        headers = ['CÓDIGO', 'NOMBRE', 'TIPO', 'DESCRIPCIÓN']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        
        # Datos de la BD
        programs = Program.objects.filter(active=True).order_by('codeProgram')
        for row_idx, program in enumerate(programs, 2):
            ws.cell(row=row_idx, column=1, value=program.codeProgram)
            ws.cell(row=row_idx, column=2, value=program.name)
            ws.cell(row=row_idx, column=3, value=program.typeProgram)
            ws.cell(row=row_idx, column=4, value=program.description)
        
        self._auto_adjust_columns(ws)

    def _create_fichas_sheet(self, workbook):
        """Crea una hoja con las fichas disponibles"""
        ws = workbook.create_sheet("Fichas")
        
        # Encabezado
        headers = ['NÚMERO FICHA', 'PROGRAMA', 'CÓDIGO PROGRAMA']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, self.header_style)
        
        # Datos de la BD
        fichas = Ficha.objects.filter(active=True).select_related('program').order_by('file_number')
        for row_idx, ficha in enumerate(fichas, 2):
            ws.cell(row=row_idx, column=1, value=ficha.file_number)
            ws.cell(row=row_idx, column=2, value=ficha.program.name)
            ws.cell(row=row_idx, column=3, value=ficha.program.codeProgram)
        
        self._auto_adjust_columns(ws)

    def _create_identification_types_sheet(self, workbook):
        """Crea una hoja con los tipos de identificación"""
        ws = workbook.create_sheet("Tipos de Identificación")
        
        # Encabezado
        cell = ws.cell(row=1, column=1, value='TIPOS DE IDENTIFICACIÓN')
        self._apply_style(cell, self.header_style)
        
        # Tipos de identificación comunes
        id_types = [
            'CC',  # Cédula de Ciudadanía
            'TI',  # Tarjeta de Identidad
            'CE',  # Cédula de Extranjería
            'PP',  # Pasaporte
            'PEP'  # Permiso Especial de Permanencia
        ]
        
        for row_idx, id_type in enumerate(id_types, 2):
            ws.cell(row=row_idx, column=1, value=id_type)
        
        self._auto_adjust_columns(ws)

    def _create_instructor_instructions_sheet(self, workbook):
        """Crea una hoja con instrucciones para instructores"""
        ws = workbook.create_sheet("Instrucciones")
        
        instructions = [
            "INSTRUCCIONES PARA EL REGISTRO MASIVO DE INSTRUCTORES",
            "",
            "CAMPOS OBLIGATORIOS (marcados con *):",
            "• Primer Nombre, Primer Apellido: Nombres completos",
            "• Tipo y Número de Identificación: Usar lista desplegable para tipo",
            "• Teléfono: Número de contacto (incluir código de país si es internacional)",
            "• Email Institucional: Debe terminar en @sena.edu.co",
            "• Tipo de Contrato: Seleccionar de la lista desplegable",
            "• Fechas de Contrato: Formato YYYY-MM-DD (Ej: 2024-01-15)",
            "• Área de Conocimiento: Seleccionar de la lista desplegable",
            "• Contraseña Temporal: Mínimo 8 caracteres",
            "",
            "LISTAS DESPLEGABLES DISPONIBLES:",
            "• Tipo de Identificación: Haz clic en la celda para ver opciones",
            "• Tipo de Contrato: Haz clic en la celda para ver opciones",
            "• Área de Conocimiento: Haz clic en la celda para ver opciones",
            "",
            "FORMATOS IMPORTANTES:",
            "• Fechas: YYYY-MM-DD",
            "• Email: usuario@sena.edu.co",
            "• Teléfono: Solo números y signos + -",
            "",
            "NOTAS:",
            "• No modificar los nombres de las columnas",
            "• Los campos opcionales pueden quedar vacíos",
            "• Usar las listas desplegables para evitar errores de escritura",
            "• La contraseña será temporal y el usuario deberá cambiarla al primer login"
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:  # Título
                self._apply_style(cell, self.header_style)
            elif instruction.startswith("•"):  # Viñetas
                cell.font = Font(italic=True)
        
        self._auto_adjust_columns(ws)

    def _create_aprendiz_instructions_sheet(self, workbook):
        """Crea una hoja con instrucciones para aprendices"""
        ws = workbook.create_sheet("Instrucciones")
        
        instructions = [
            "INSTRUCCIONES PARA EL REGISTRO MASIVO DE APRENDICES",
            "",
            "CAMPOS OBLIGATORIOS (marcados con *):",
            "• Primer Nombre, Primer Apellido: Nombres completos",
            "• Tipo y Número de Identificación: Usar lista desplegable para tipo",
            "• Teléfono: Número de contacto",
            "• Email Institucional: Debe terminar en @soy.sena.edu.co",
            "• Código Programa: Seleccionar de la lista desplegable",
            "• Número Ficha: Seleccionar de la lista desplegable",
            "• Contraseña Temporal: Mínimo 8 caracteres",
            "",
            "LISTAS DESPLEGABLES DISPONIBLES:",
            "• Tipo de Identificación: Haz clic en la celda para ver opciones",
            "• Código Programa: Haz clic en la celda para ver opciones",
            "• Número Ficha: Haz clic en la celda para ver opciones",
            "",
            "FORMATOS IMPORTANTES:",
            "• Email: usuario@soy.sena.edu.co (diferente a instructores)",
            "• Teléfono: Solo números y signos + -",
            "",
            "NOTAS:",
            "• No modificar los nombres de las columnas",
            "• Los campos opcionales pueden quedar vacíos",
            "• Usar las listas desplegables para evitar errores de escritura",
            "• La contraseña será temporal y el usuario deberá cambiarla al primer login",
            "• Los aprendices deben usar el dominio @soy.sena.edu.co"
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=row_idx, column=1, value=instruction)
            if row_idx == 1:  # Título
                self._apply_style(cell, self.header_style)
            elif instruction.startswith("•"):  # Viñetas
                cell.font = Font(italic=True)
        
        self._auto_adjust_columns(ws)

    def _save_workbook_to_response(self, workbook, filename):
        """Guarda el workbook en un HttpResponse para descarga"""
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    def _add_instructor_data_validations(self, worksheet):
        """Agrega validaciones de datos (listas desplegables) para la plantilla de instructores"""
        
        # Validación para Tipo de Identificación (columna E)
        id_types = ['CC', 'TI', 'CE', 'PP', 'PEP']
        self._add_data_validation(worksheet, 'E', id_types)
        
        # Validación para Tipo de Contrato (columna I)
        contract_types = ['Planta', 'Contratista', 'Temporal', 'Prestación de Servicios', 'Cátedra']
        self._add_data_validation(worksheet, 'I', contract_types)
        
        # Validación para Área de Conocimiento (columna L)
        try:
            knowledge_areas = list(KnowledgeArea.objects.filter(active=True).values_list('name', flat=True))
            if knowledge_areas:
                self._add_data_validation(worksheet, 'L', knowledge_areas)
        except Exception as e:
            print(f"Error obteniendo áreas de conocimiento: {e}")

    def _add_aprendiz_data_validations(self, worksheet):
        """Agrega validaciones de datos (listas desplegables) para la plantilla de aprendices"""
        
        # Validación para Tipo de Identificación (columna E)
        id_types = ['CC', 'TI', 'CE', 'PP', 'PEP']
        self._add_data_validation(worksheet, 'E', id_types)
        
        # Validación para Código de Programa (columna I)
        try:
            program_codes = list(Program.objects.filter(active=True).values_list('codeProgram', flat=True))
            if program_codes:
                self._add_data_validation(worksheet, 'I', program_codes)
        except Exception as e:
            print(f"Error obteniendo códigos de programa: {e}")
        
        # Validación para Número de Ficha (columna J)
        try:
            ficha_numbers = list(Ficha.objects.filter(active=True).values_list('file_number', flat=True))
            if ficha_numbers:
                self._add_data_validation(worksheet, 'J', ficha_numbers)
        except Exception as e:
            print(f"Error obteniendo números de ficha: {e}")